# -*- coding: utf-8  -*-
"""
Robot which will does substitutions of tags within wiki page content with external or
other wiki text data. Like dynamic text updating.

Look at http://de.wikipedia.org/wiki/Benutzer:DrTrigon/Benutzer:DrTrigonBot/config.css
for 'postproc' example code.

Look at https://wiki.toolserver.org/view/Mail how to setup mail handling. The
following code was used in file '$HOME/.forward+subster':
--- --- --- --- --- --- --- --- --- ---
> ~/data/subster/mail_inbox
--- --- --- --- --- --- --- --- --- ---
in order to enable mail (mbox style) storage in given file for address:
drtrigon+subster@toolserver.org

Other scripts and tools related to this bot are:
- subster_irc.py            IRC Robot
- substersim.py             Subster simulation panel
- subster_mail_queue.py     Subster mail queue

The following parameters are supported:

&params;

All other parameters will be ignored.

Syntax example:
    python subster.py
        Default operating mode.

    python subster.py -lang:en
        Run bot on another site language than configured as default. E.g. 'en'.

    python subster.py -family:meta -lang:
    python subster.py -family:wikidata -lang:wikidata
        Run bot on another site family and language than configured as default.
        E.g. 'meta' or 'wikidata'.

    python subster_irc.py
        Default operating mode for IRC Robot. The IRC bot uses this script as
        subclass.
"""
## @package subster
#  @brief   Dynamic Text Substitutions Robot
#
#  @copyright Dr. Trigon, 2009-2012
#
#  @section FRAMEWORK
#
#  Python wikipedia robot framework, DrTrigonBot.
#  @see http://pywikipediabot.sourceforge.net/
#  @see http://de.wikipedia.org/wiki/Benutzer:DrTrigonBot
#
#  External code / other modules used are listed here.
#  @see https://bitbucket.org/ericgazoni/openpyxl/wiki/Home
#  @see http://pypi.python.org/pypi/crontab/.11
#
#  @section LICENSE
#
#  Distributed under the terms of the MIT license.
#  @see http://de.wikipedia.org/wiki/MIT-Lizenz
#
__version__ = '$Id: 01ab3a69ad9ba660595fca72d67fe51a20bb6766 $'
#


import re, sys, os, string, time, copy
import difflib, traceback
import StringIO, zipfile, csv
import mailbox, mimetypes, datetime, email.utils
import logging
import ast
import shelve, pprint

import pagegenerators, basic
# Splitting the bot into library parts
import wikipedia as pywikibot
from pywikibot import i18n
from pywikibot.comms import http
import catlib
import externals                            # check for and install needed
externals.check_setup('BeautifulSoup.py')   #  'externals'
externals.check_setup('crontab')            #
externals.check_setup('odf')                #
externals.check_setup('openpyxl')           #
import BeautifulSoup
import openpyxl.reader.excel
import crontab


# TODO: think about what config to move to 'subster-config.css' (per wiki)
#       e.g. 'VerboseMessage', 'data_VerboseMessage', ...
bot_config = {
        # unicode values
             'BotName':     pywikibot.config.usernames[pywikibot.config.family][pywikibot.config.mylang],
        'TemplateName':     u'User:DrTrigonBot/Subster',    # or 'template' for 'Flagged Revisions'
   'data_PropertyId':       u'370',                         # default: Sandbox-String (P370)

        'ErrorTemplate':    u'<b>SubsterBot Exception in "%s" (%s)</b>\n<pre>%s</pre>',
        'VerboseMessage':   u'<noinclude>\n----\n%s\n</noinclude>', # DRTRIGON-116, DRTRIGON-132
   'data_VerboseMessage':   u'<onlyinclude>{{#switch: {{{1|}}}\n'
                            u'|          error = %(error)s\n'
                            u'|error-traceback = %(error-traceback)s\n'
                            u'}}</onlyinclude>',

        # important to use a '.css' page here, since it HAS TO BE protected to
        # prevent malicious code injection !
        'ConfCSSpostproc':  u'User:DrTrigon/DrTrigonBot/subster-postproc.css',
        'ConfCSSconfig':    u'User:DrTrigon/DrTrigonBot/subster-config.css',

        'CodeTemplate':     u'\n%s(DATA, *args)\n',
        'CRONMaxDelay':     1*24*60*60,       # bot runs daily

        # regex values
        'var_regex_str':    u'<!--SUBSTER-%(var1)s-->%(cont)s<!--SUBSTER-%(var2)s-->',

        'mbox_file':        'mail_inbox',    # "drtrigon+subster@toolserver.org"
        'data_path':        '../data/subster',

        # bot paramater/options
        'param_default':    {
            'url':             '',
            'regex':           '',
            'value':           '',
            'count':           '0',
            #'postproc':        '("","")',
            'postproc':        '(\'\', \'\')',
            'beautifulsoup':   'False',         # DRTRIGON-88
            'expandtemplates': 'False',         # DRTRIGON-93 (with 'wiki://')
            'simple':          '',              # DRTRIGON-85
            'zip':             'False',
            'xlsx':            '',              #
            'ods':             '',              #
            # may be 'hours' have to be added too (e.g. for 'ar')
            'cron':            '',              # DRTRIGON-102
            'verbose':         'True',          # DRTRIGON-132 (else see logs)
            #'djvu': ... u"djvused -e 'n' \"%s\"" ... djvutext.py
            #'pdf': ... u"pdftotext" or python module
            #'imageocr', 'swfocr', ...
        },

        # this is a system parameter and should not be changed! (copy.deepcopy)
        'EditFlags':        {'minorEdit': True, 'botflag': True},
}


class SubsterBot(basic.AutoBasicBot):
    '''
    Robot which will does substitutions of tags within wiki page content with external or
    other wiki text data. Like dynamic text updating.
    '''

    _var_regex_str = bot_config['var_regex_str']%{'var1':'%(var)s','var2':'%(var)s','cont':'%(cont)s'}
    _BS_regex_str  = bot_config['var_regex_str']%{'var1':'%(var1)s','var2':'%(var2)sBS:/','cont':'%(cont)s'}

    # -template and subst-tag handling taken from MerlBot
    # -this bot could also be runned on any local wiki with an anacron-job

    def __init__(self):
        '''Constructor of SubsterBot(), initialize needed vars.'''

        pywikibot.output(u'\03{lightgreen}* Initialization of bot:\03{default}')

        basic.AutoBasicBot.__init__(self)

        # modification of timezone to be in sync with wiki
        os.environ['TZ'] = 'Europe/Amsterdam'
        if hasattr(time, "tzset"):
            time.tzset()
            pywikibot.output(u'Setting process TimeZone (TZ): %s' % str(time.tzname))    # ('CET', 'CEST')
        else:
            # e.g. windows doesn't have that attribute
            pywikibot.warning(u'This operating system has NO SUPPORT for setting TimeZone by code! Before running this script, please set the TimeZone manually to one approriate for use with the Wikipedia language and region you intend to.')

        # init constants
        self._bot_config = bot_config
        # convert e.g. namespaces to corret language
        self._bot_config['TemplateName'] = pywikibot.Page(self.site, self._bot_config['TemplateName']).title()
        self._template_regex = re.compile('\{\{' + self._bot_config['TemplateName'] + '(.*?)\}\}', re.S)
        if self.site.is_data_repository():
            self._bot_config['VerboseMessage'] = self._bot_config['data_VerboseMessage']

        # init constants
        self._userListPage        = pywikibot.Page(self.site, self._bot_config['TemplateName'])
        self._ConfCSSpostprocPage = pywikibot.Page(self.site, self._bot_config['ConfCSSpostproc'])
        self._ConfCSSconfigPage   = pywikibot.Page(self.site, self._bot_config['ConfCSSconfig'])
        self.pagegen = pagegenerators.ReferringPageGenerator(self._userListPage, onlyTemplateInclusion=True)
        self._code   = self._ConfCSSpostprocPage.get()
        pywikibot.output(u'Imported postproc %s rev %s from %s' %\
          ((self._ConfCSSpostprocPage.title(asLink=True),) + self._ConfCSSpostprocPage.getVersionHistory(revCount=1)[0][:2]) )
        self._flagenable = {}
        if self._ConfCSSconfigPage.exists():
            exec(self._ConfCSSconfigPage.get())    # with variable: bot_config_wiki
            self._flagenable = bot_config_wiki['flagenable']
            pywikibot.output(u'Imported config %s rev %s from %s' %\
              ((self._ConfCSSconfigPage.title(asLink=True),) + self._ConfCSSconfigPage.getVersionHistory(revCount=1)[0][:2]) )

    def run(self, sim=False, msg=None, EditFlags=bot_config['EditFlags']):
        '''Run SubsterBot().'''

        pywikibot.output(u'\03{lightgreen}* Processing Template Backlink List:\03{default}')

        if sim:    self.pagegen = ['dummy']

        for page in self.pagegen:
            # setup source to get data from
            if sim:
                content = sim['content']
                params = [ sim ]
            else:
                pywikibot.output(u'Getting page "%s" via API from %s...'
                                 % (page.title(asLink=True), self.site))

                # get page content and operating mode
                content = self.load(page)
                params = self.loadTemplates(page, self._bot_config['TemplateName'],
                                            default=self._bot_config['param_default'])

            if not params: continue

            (substed_content, substed_tags) = self.subContent(content, params)

            # output result to page or return directly
            if sim:
                return substed_content
            else:
                # if changed, write!
                if (substed_content != content):
                #if substed_tags:
                    self.outputContentDiff(content, substed_content)

                    head = i18n.twtranslate(self.site.lang,
                                            'thirdparty-drtrigonbot-sum_disc-summary-head')
                    if msg is None:
                        msg = i18n.twtranslate(self.site.lang,
                                               'thirdparty-drtrigonbot-subster-summary-mod')
                    flags = copy.deepcopy(EditFlags)
                    if page.title() in self._flagenable:
                        flags.update( self._flagenable[page.title()] )
                    pywikibot.output(u'Flags used for writing: %s' % flags)
                    self.save( page, substed_content,
                               (head + u' ' + msg) % {'tags':", ".join(substed_tags)},
                               **flags )

                    # DRTRIGON-130: data repository (wikidata) output to items
                    if self.site.is_data_repository():
                        data = self.data_convertContent(substed_content)
                        self.data_save(page, data)
                else:
                    pywikibot.output(u'NOTHING TO DO!')

    def subContent(self, content, params):
        """Substitute the tags in content according to params.

           @param content: Content with tags to substitute.
           @type  content: string
           @param params: Params with data how to substitute tags.
           @type  params: dict

           Returns a tuple containig the new content with tags
           substituted and a list of those tags.
        """

        #md_val_tag = u'%s-META-%s'
        md_val_tag = u'META-%s-%s'

        substed_content = content
        substed_tags = []  # DRTRIGON-73

        for item in params:
            # 1st stage: main/general content substitution
            # 1.) - 5.) subst templates
            metadata = { 'bot-error':           unicode(False),
                         'bot-error-traceback': u'', }  # DRTRIGON-132
            try:
                (substed_content, tags, md) = self.subTemplate(substed_content, item)
                substed_tags += tags
                metadata.update(md)

                # DRTRIGON-132; metadata append IFF other data/content changed
                # (can change all the time, but MUST NOT trigger a page save/change!)
                if not tags:
                    metadata = {}
            except:
                exc_info = sys.exc_info()
                result = u''.join(traceback.format_exception(exc_info[0], exc_info[1], exc_info[2]))

                # DRTRIGON-132; metadata append IFF exception raised
                # (this metadata HAVE TO trigger a change because of error!)
                metadata['bot-error'] = unicode(True)
                metadata['bot-error-traceback'] = self._bot_config['ErrorTemplate'] %\
                                     ( item['value'],
                                       pywikibot.Timestamp.now().isoformat(' '),
                                       result.strip() )

                # VerboseMode: IFF no 'bot-error-traceback' metadata tag present on
                # page, append it in order not to loose error info (single exception)
                value = md_val_tag % (item['value'], 'bot-error-traceback')
                tags = self.subTag(substed_content, value)[1]
                if ast.literal_eval(item['verbose']) and (value not in tags):
                    substed_content += self._bot_config['VerboseMessage'] %\
                      (self._var_regex_str % {'var': value, 'cont': u''})

            # 2nd stage: conditional metadata substitution (DRTRIGON-132)
            # (IFF content changed, exception raised, ...)
            for data in metadata:
                value = md_val_tag % (item['value'], data)
                (substed_content, tags) = self.subTag(substed_content, value, metadata[data], 0)
                substed_tags += tags
                #substed_tags.append( u'>error:%s<' % item['value'] )

        return (substed_content, substed_tags)

    def subTemplate(self, content, param):
        """Substitute the template tags in content according to param.

           @param content: Content with tags to substitute.
           @type  content: string
           @param param: Param with data how to substitute tags.
           @type  param: dict

           Returns a tuple containig the new content with tags
           substituted and a list of those tags.
        """

        substed_tags = []  # DRTRIGON-73
        metadata     = {'mw-signature': u'~~~~',
                        'mw-timestamp': u'~~~~~',}  # DRTRIGON-132

        # 0.2.) check for 'simple' mode and get additional params
        if param['simple']:
            p = self.site.getExpandedString(param['simple'])
            param.update(pywikibot.extract_templates_and_params(p)[0][1])

        # 0.5.) check cron/date
        if param['cron']:
            # [min] [hour] [day of month] [month] [day of week]
            # (date supported only, thus [min] and [hour] dropped)
            if not (param['cron'][0] == '@'):
                param['cron'] = '* * ' + param['cron']
            entry = crontab.CronTab(param['cron'])
            # find the delay from midnight (does not return 0.0 - but next)
            delay = entry.next(datetime.datetime.now().replace(hour=0,
                                                               minute=0,
                                                               second=0,
                                                               microsecond=0)- \
                               datetime.timedelta(microseconds=1))

            pywikibot.output(u'CRON delay for execution: %.3f (<= %i)'
                             % (delay, self._bot_config['CRONMaxDelay']))

            if not (delay <= self._bot_config['CRONMaxDelay']):
                return (content, substed_tags, metadata)

        # 1.) getUrl or wiki text
        # (security: check url not to point to a local file on the server,
        #  e.g. 'file://' - same as used in xsalt.py)
        secure = False
        for item in [u'http://', u'https://',
                     u'mail://', u'local://', u'wiki://']:
            secure = secure or (param['url'][:len(item)] == item)
        param['zip'] = ast.literal_eval(param['zip'])
        if not secure:
            return (content, substed_tags, metadata)
        if   (param['url'][:7] == u'wiki://'):
            url = param['url'][7:].strip('[]')              # enable wiki-links
            if ast.literal_eval(param['expandtemplates']):  # DRTRIGON-93 (only with 'wiki://')
                external_buffer = pywikibot.Page(self.site,
                                                 url).get(expandtemplates=True)
            else:
                external_buffer = self.load( pywikibot.Page(self.site, url) )
        elif (param['url'][:7] == u'mail://'):              # DRTRIGON-101
            url = param['url'].replace(u'{{@}}', u'@')     # e.g. nlwiki
            mbox = SubsterMailbox(
              pywikibot.config.datafilepath(self._bot_config['data_path'],
                                            self._bot_config['mbox_file'], ''))
            external_buffer = mbox.find_data(url)
            mbox.close()
        elif (param['url'][:8] == u'local://'):             # DRTRIGON-131
            if (param['url'][8:] == u'cache/state_bots'):
                # filename hard-coded
                d = shelve.open(pywikibot.config.datafilepath('cache',
                                                              'state_bots'))
                external_buffer = pprint.pformat(
                    ast.literal_eval(pprint.pformat(d)))
                d.close()
            else:
                external_buffer = u'n/a'
        else:
            # consider using 'expires', 'last-modified', 'etag' in order to
            # make the updating data requests more efficient! use those stored
            # on page, if the user placed them, else use the conventional mode.
            # http://www.diveintopython.net/http_web_services/etags.html
            f_url, external_buffer = http.request(self.site, param['url'],
                                                  no_hostname = True,
                                                  back_response = True)
            headers = f_url.headers # same like 'f_url.info()'
            #if param['zip']:
            if ('text/' not in headers['content-type']):
                pywikibot.output(u'Source is of non-text content-type, '
                                 u'using raw data instead.')
                external_buffer = f_url.read()
            del f_url               # free some memory (no need to keep copy)

            for h in ['content-length', 'date', 'last-modified', 'expires']:
                if h in headers:
                    metadata['url-%s' % h] = headers[h]

        # some intermediate processing (unzip, xlsx2csv, ...)
        if param['zip']:    # 'application/zip', ...
            fileno          = 0 if (param['zip'] == True) else (param['zip']-1)
            external_buffer = self.unzip(external_buffer, fileno)
        if param['xlsx']:   # 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            external_buffer = self.xlsx2csv(external_buffer, param['xlsx'])
        if param['ods']:    # 'application/vnd.oasis.opendocument.spreadsheet'
            external_buffer = self.ods2csv(external_buffer, param['ods'])

        if not ast.literal_eval(param['beautifulsoup']):    # DRTRIGON-88
            # 2.) regexp
            #for subitem in param['regex']:
            subitem = param['regex']
            regex = re.compile(subitem, re.S | re.I)

            # 3.) subst in content
            external_data = regex.search(external_buffer)

            external_data_dict = {}
            if external_data:    # not None
                external_data = external_data.groups()

                pywikibot.output(u'Groups found by regex: %i'
                                 % len(external_data))

                # DRTRIGON-114: Support for named groups in regexs
                if regex.groupindex:
                    for item in regex.groupindex:
                        external_data_dict[u'%s-%s' % (param['value'], item)] = external_data[regex.groupindex[item]-1]
                elif (len(external_data) == 1):
                    external_data_dict = {param['value']: external_data[0]}
                else:
                    external_data_dict = {param['value']: str(external_data)}
            pywikibot.debug( str(external_data_dict) )

            param['postproc'] = eval(param['postproc'])
            # should be secured as given below, but needs code changes in wiki too
            #param['postproc'] = ast.literal_eval(param['postproc'])
            for value in external_data_dict:
                external_data = external_data_dict[value]

                # 4.) postprocessing
                func  = param['postproc'][0]    # needed by exec call of self._code
                DATA  = [ external_data ]       #
                args  = param['postproc'][1:]   #
                scope = {}                      # (scope to run in)
                scope.update( locals() )        # (add DATA, *args, ...)
                scope.update( globals() )       # (add imports and else)
                if func:
                    exec(self._code + (self._bot_config['CodeTemplate'] % func), scope, scope)
                    external_data = DATA[0]
                pywikibot.debug( external_data )

                # 5.) subst content
                (content, tags) = self.subTag(content, value, external_data, int(param['count']))
                substed_tags += tags
        else:
            # DRTRIGON-105: Support for multiple BS template configurations
            value = param['value']
            if value:
                value += u'-'

            # DRTRIGON-88: Enable Beautiful Soup power for Subster
            BS_tags = self.get_BS_regex(value).findall(content)

            pywikibot.output(u'BeautifulSoup tags found by regex: %i' % len(BS_tags))

            prev_content = content

            BS = BeautifulSoup.BeautifulSoup(external_buffer)
            for item in BS_tags:
                external_data = eval('BS.%s' % item[1])
                external_data = self._BS_regex_str%{'var1':value+'BS:'+item[1],'var2':value,'cont':external_data}
                content = content.replace(item[0], external_data, 1)

            if (content != prev_content):
                substed_tags.append(value+'BS')

        metadata['bot-timestamp'] = pywikibot.Timestamp.now().isoformat(' ')

        return (content, substed_tags, metadata)

    def subTag(self, content, value, external_data=u'~~~~', count=1):
        """Substitute one single tag (of a template) in content.

           Can also be (ab)used to check for presence of a tag.
        """
        substed_tags = []

        # 5.) subst content
        prev_content = content
        var_regex = self.get_var_regex(value)
        content = var_regex.sub((self._var_regex_str%{'var':value,'cont':external_data}), content, count)
        if (content != prev_content):
            substed_tags.append(value)

        return (content, substed_tags)

    def outputContentDiff(self, content, substed_content):
        """Outputs the diff between the original and the new content.

           @param content: Original content.
           @type  content: string
           @param substed_content: New content.
           @type  substed_content: string

           Returns nothing, but outputs/prints the diff.
        """
        diff = difflib.Differ().compare(content.splitlines(1), substed_content.splitlines(1))
        diff = [ line for line in diff if line[0].strip() ]
        pywikibot.output(u'Diff:')
        pywikibot.output(u'--- ' * 15)
        pywikibot.output(u''.join(diff))
        pywikibot.output(u'--- ' * 15)

    def data_convertContent(self, substed_content):
        """Converts the substed content to Wikidata format in order to save.

           Template page format (adopted from #switch):
             <pre>
             | key1 = value1
             | key2 = value2
             ...
             </pre>
           every entry has to start with a '|' and contain a '=', the entries
           have to be embedded into pre-tags (entries may share the same line)

           @param substed_content: New/Changed content (including tags).
           @type  substed_content: string
           
           Returns the extracted and converted data.
        """
        # DRTRIGON-130: convert talk page result to wikidata(base)
        data = u'\n'.join(re.findall('<pre>(.*?)</pre>', substed_content, 
                                     re.S | re.I))
        data = self.get_var_regex('.*?', '(.*?)').sub('\g<1>', data)
        res = {}
        for line in data.split(u'|'):
            line = line.strip().split(u'=', 1)
            if len(line) != 2:
                continue
            res[line[0].strip()] = line[1].strip()

        return res

    def data_save(self, page, data):
        """Stores the content to Wikidata.

           @param page: Page containing template.
           @type  page: page object
           @param data: New content.
           @type  data: dict

           Returns nothing, but stores the changed content to linked labels.
        """
        # DRTRIGON-130: check for changes and then write/change/set values
        datapage = pywikibot.DataPage(self.site, page.title())
        dataitem = u'%s:%s' % (self._bot_config['BotName'], datapage.title().split(u':')[1])
        links  = [ {u'aliases': [u'%s:%s' % (dataitem, p.sortkeyprefix)],
                    u'id': p.toggleTalkPage().title().lower(),}
                   for p in catlib.Category(self.site, dataitem).articles() ]
        links += datapage.searchentities(dataitem)

        for element in links:
            propid = int(self._bot_config['data_PropertyId'])
            el = element[u'aliases'][0].split(u':')
            item = el[2]
            if item not in data:
                pywikibot.output(u'Value "%s" not found.' % (item,))
                data[item] = u'%s: N/A' % self._bot_config['BotName']
            if len(el) > 3:
                propid = int(el[3])

            dataoutpage = pywikibot.DataPage(self.site, element['id'])

            # check for changes and then write/change/set values
            summary = u'Bot: update data because of configuration on %s.' % page.title(asLink=True)
            buf = dataoutpage.get()
            claim = [ claim for claim in buf[u'claims'] if (claim['m'][1] == propid) ]
            # TODO: does this check (if) work with multiple claims per property?
            if (not claim) or (claim[0]['m'][3] != data[item]):
                pywikibot.output(u'%s in %s changed to "%s"' %\
                    (element[u'aliases'][0], dataoutpage.title(asLink=True), data[item]))
                dataoutpage.editclaim(u'p%s' % propid, data[item],
                                      refs={"p%s" % propid:
                                          [{"snaktype":  "value",
                                            "property":  "p%s" % propid,
                                            "datavalue": {u'type':  u'string', 
                                                          u'value': datapage.title()}},
                                           {"snaktype":  "value",
                                            "property":  "p585",    # point in time
                                            #"property":  "p578",    # Sandbox-TimeValue
                                            "datavalue": {u'type':  u'time',
                                                          u'value': {u'after':         0, 
                                                                     u'precision':     11, 
                                                                     u'time':          (u'+0000000%sZ' % pywikibot.Timestamp.now().isoformat().split('.')[0]), 
                                                                     u'timezone':      0, 
                                                                     u'calendarmodel': u'http://www.wikidata.org/entity/Q1985727', 
                                                                     u'before':        0}}},]},
                                      comment=summary)
        #print data['timestampFIDE'], pywikibot.Timestamp.now().isoformat()

    def get_var_regex(self, var, cont='.*?'):
        """Get regex used/needed to find the tags to replace.

           @param var: The tag/variable name.
           @type  var: string
           @param cont: The content/value of the variable.
           @type  cont: string

           Return the according (and compiled) regex object.
        """
        return re.compile((self._var_regex_str%{'var':var,'cont':cont}), re.S | re.I)

    def get_BS_regex(self, var, cont='(.*?)'):
        """Get regex used/needed to find the BS tags to replace.

           @param var: The tag/variable name.
           @type  var: string
           @param cont: The content/value of the variable.
           @type  cont: string

           Return the according (and compiled) regex object.
        """
        return re.compile(u'(' + self._BS_regex_str%{'var1':var+'BS:(.*?)','var2':var,'cont':cont} + u')')

    def unzip(self, external_buffer, i):
        """Convert zip data to plain format.
        """

        zip_buffer = zipfile.ZipFile(StringIO.StringIO(external_buffer))
        data_file  = zip_buffer.namelist()[i]
        external_buffer = zip_buffer.open(data_file).read().decode('latin-1')

        return external_buffer

    def xlsx2csv(self, external_buffer, sheet):
        """Convert xlsx (EXCEL) data to csv format.
        """

        wb = openpyxl.reader.excel.load_workbook(StringIO.StringIO(external_buffer), use_iterators = True)

        sheet_ranges = wb.get_sheet_by_name(name = sheet)

        output = StringIO.StringIO()
        spamWriter = csv.writer(output)

        for row in sheet_ranges.iter_rows(): # it brings a new method: iter_rows()
            spamWriter.writerow([ cell.internal_value for cell in row ])

        external_buffer = output.getvalue()
        output.close()

        return external_buffer

    def ods2csv(self, external_buffer, sheet):
        """Convert ods (Open/Libre Office) data to csv format.
        """
        # http://www.mail-archive.com/python-list@python.org/msg209447.html

        import odf
        from odf import opendocument, table, teletype

        doc = odf.opendocument.load(StringIO.StringIO(external_buffer))

        output = StringIO.StringIO()
        spamWriter = csv.writer(output)

        for sheet in doc.getElementsByType(odf.table.Table):
            if not (sheet.getAttribute('name') == sheet):
                continue
            for row in sheet.getElementsByType(odf.table.TableRow):
                spamWriter.writerow([ odf.teletype.extractText(cell).encode('utf-8')
                                      for cell in row.getElementsByType(odf.table.TableCell) ])

        external_buffer = output.getvalue()
        output.close()

        return external_buffer


class SubsterMailbox(mailbox.mbox):
    def __init__(self, mbox_file):
        mailbox.mbox.__init__(self, mbox_file)
        self.lock()

        self.remove_duplicates()

    def remove_duplicates(self):
        """Find mails with same 'From' (sender) and remove all
        except the most recent one.
        """

        unique = {}
        remove = []
        for i, message in enumerate(self):
            sender   = message['from']       # Could possibly be None.
            timestmp = message['date']       # Could possibly be None.

            timestmp = time.mktime( email.utils.parsedate(timestmp) )
            timestmp = datetime.datetime.fromtimestamp( timestmp )

            if sender in unique:
                (j, timestmp_j) = unique[sender]

                if (timestmp >= timestmp_j):
                    remove.append( j )
                else:
                    remove.append( i )
            else:
                unique[sender] = (i, timestmp)

        remove.reverse()
        for i in remove:
            self.remove(i)

        self.flush()
        #self.close()

        if remove:
            pywikibot.output('Removed %i depreciated email data source(s).' % len(remove))

    def find_data(self, url):
        """Find mail according to given 'From' (sender).
        """

        url = (url[:7], ) + tuple(url[7:].split('/'))
        content = []

        for i, message in enumerate(self):
            sender   = message['from']       # Could possibly be None.
            subject  = message['subject']    # Could possibly be None.
            timestmp = message['date']       # Could possibly be None.

            if sender and url[1] in sender:
                # data found
                pywikibot.output('Found email data source:')
                pywikibot.output('%i / %s / %s / %s' % (i, sender, subject, timestmp))

                full = (url[2] == 'attachment-full')
                ind  = 0    # default; ignore attachement index
                if   (url[2] == 'all'):
                    content = [ message.as_string(True) ]
                elif (url[2] == 'attachment') or full:
                    if len(url) > 3:
                        ind = int(url[3])   # combine 'ind' with 'full=True'...?
                    counter = 1
                    content = []
                    for part in message.walk():
                        # multipart/* are just containers
                        if part.get_content_maintype() == 'multipart':
                            continue
                        # Applications should really sanitize the given filename so that an
                        # email message can't be used to overwrite important files
                        filename = part.get_filename()
                        if filename or full:
                            if not filename:
                                ext = mimetypes.guess_extension(part.get_content_type())
                                if not ext:
                                    # Use a generic bag-of-bits extension
                                    ext = '.bin'
                                filename = 'part-%03d%s' % (counter, ext)

                            content += [ part.get_payload(decode=True) ]
                            pywikibot.output('Found attachment # %i: "%s"' % (counter, filename))

                            if counter == ind:
                                return content[-1]

                            counter += 1

                            if (not full) and (not ind):
                                break

                    break

        return string.join(content)


def main():
    args = pywikibot.handleArgs()
    bot  = SubsterBot()   # for several user's, but what about complete automation (continous running...)
    for arg in args:
        pywikibot.showHelp()
        return
    try:
        bot.run()
    except KeyboardInterrupt:
        pywikibot.output('\nQuitting program...')

if __name__ == "__main__":
    try:
        main()
    finally:
        pywikibot.stopme()

